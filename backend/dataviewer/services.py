from django.conf import settings
import openpyxl
from typing import List, Dict, Any
import json


class ExcelDataService:
    """Excel 데이터를 읽고 처리하는 서비스"""

    @staticmethod
    def get_metadata() -> List[Dict[str, Any]]:
        """meta 시트에서 컬럼 메타데이터를 가져옴"""
        workbook = openpyxl.load_workbook(settings.DATA_FILE_PATH, read_only=True)

        if 'meta' not in workbook.sheetnames:
            workbook.close()
            return []

        meta_sheet = workbook['meta']
        headers = []
        metadata = []

        for idx, row in enumerate(meta_sheet.iter_rows(values_only=True)):
            if idx == 0:
                headers = list(row)
            else:
                if any(row):  # 빈 행이 아닌 경우
                    metadata.append(dict(zip(headers, row)))

        workbook.close()
        return metadata

    @staticmethod
    def get_data(page: int = 1, page_size: int = 50) -> Dict[str, Any]:
        """infos 시트에서 데이터를 페이징하여 가져옴"""
        workbook = openpyxl.load_workbook(settings.DATA_FILE_PATH, read_only=True)

        if 'infos' not in workbook.sheetnames:
            workbook.close()
            return {'data': [], 'total': 0, 'page': page, 'page_size': page_size}

        infos_sheet = workbook['infos']
        headers = []
        all_data = []

        for idx, row in enumerate(infos_sheet.iter_rows(values_only=True)):
            if idx == 0:
                headers = list(row)
            else:
                if any(row):  # 빈 행이 아닌 경우
                    all_data.append(dict(zip(headers, row)))

        workbook.close()

        # 페이징 처리
        total = len(all_data)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_data = all_data[start_idx:end_idx]

        return {
            'data': paginated_data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }

    @staticmethod
    def update_data(row_index: int, column_name: str, value: Any) -> bool:
        """infos 시트의 특정 셀 데이터를 업데이트"""
        try:
            # read_only=False로 워크북 열기
            workbook = openpyxl.load_workbook(settings.DATA_FILE_PATH, read_only=False)

            if 'infos' not in workbook.sheetnames:
                workbook.close()
                raise ValueError("infos 시트를 찾을 수 없습니다.")

            infos_sheet = workbook['infos']

            # 헤더 찾기
            headers = []
            for cell in infos_sheet[1]:
                headers.append(cell.value)

            # 컬럼 인덱스 찾기
            if column_name not in headers:
                workbook.close()
                raise ValueError(f"컬럼 '{column_name}'을 찾을 수 없습니다.")

            col_index = headers.index(column_name) + 1  # openpyxl은 1부터 시작

            # 실제 행 번호 (헤더 제외하고 row_index번째 데이터 행)
            actual_row = row_index + 2  # 헤더(1행) + row_index

            # 값 저장 (중재활동분류는 쉼표로 구분된 문자열로 저장)
            if column_name == '중재활동분류':
                if isinstance(value, list):
                    # 배열을 "4, 5" 형식의 문자열로 변환
                    cell_value = ', '.join(map(str, value)) if value else ''
                else:
                    cell_value = value
            else:
                cell_value = value

            # 셀에 값 쓰기
            infos_sheet.cell(row=actual_row, column=col_index, value=cell_value)

            # 파일 저장
            workbook.save(settings.DATA_FILE_PATH)
            workbook.close()

            return True

        except Exception as e:
            print(f"데이터 업데이트 오류: {str(e)}")
            raise e
